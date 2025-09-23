"""
Views for Transaction Management API
"""

import logging
from datetime import timedelta
from django.utils import timezone
from django.db import transaction as db_transaction
from django.db.models import Q, Count, Avg, F
from django.shortcuts import get_object_or_404
from django.conf import settings
from rest_framework import viewsets, status, permissions, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from .models import Transaction, TransactionStatusHistory, Comment
from attachments.models import Attachment
from attachments.serializers import AttachmentSerializer
from .serializers import (
    TransactionSerializer, TransactionListSerializer, TransactionCreateSerializer,
    TransactionStatusUpdateSerializer, TransactionAssignmentSerializer,
    TransactionFilterSerializer, BulkTransactionOperationSerializer,
    TransactionStatisticsSerializer, CommentSerializer
)
from core.permissions import (
    IsActiveUser, IsAdminUser, IsEditorOrAdmin, CanManageTransactions,
    CanViewTransaction, CanModifyTransaction
)
from core.utils import (
    create_success_response, create_error_response, send_notification_email,
    create_audit_log_entry, get_client_ip
)
from core.pagination import StandardPagination, LargePagination
from core.throttling import BulkOperationThrottle

logger = logging.getLogger(__name__)


class TransactionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Transaction CRUD operations with role-based filtering
    """
    pagination_class = StandardPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = [
        'transaction_id', 'reference_number', 'client_name',
        'transaction_type', 'description', 'tags'
    ]
    ordering_fields = [
        'transaction_id', 'client_name', 'status', 'priority',
        'created_at', 'updated_at', 'due_date'
    ]
    ordering = ['-created_at']
    filterset_fields = [
        'status', 'priority', 'assigned_to',
        'client', 'created_by'
    ]
    
    def get_permissions(self):
        """
        Instantiate and return the list of permissions that this view requires.
        """
        if self.action == 'create':
            permission_classes = [IsActiveUser]
        elif self.action in ['list', 'retrieve']:
            permission_classes = [IsActiveUser]
        elif self.action in ['update', 'partial_update', 'destroy']:
            permission_classes = [IsActiveUser]  # Will check object permissions
        else:
            permission_classes = [IsActiveUser]
        
        return [permission() for permission in permission_classes]
    
    def get_serializer_class(self):
        """
        Return the appropriate serializer class based on action
        """
        if self.action == 'list':
            return TransactionListSerializer
        elif self.action == 'create':
            return TransactionCreateSerializer
        return TransactionSerializer
    
    def get_queryset(self):
        """
        Return filtered queryset based on user role
        """
        user = self.request.user

        if not user.is_authenticated:
            return Transaction.objects.none()

        # Base queryset excluding soft-deleted transactions
        queryset = Transaction.objects.filter(is_deleted=False).select_related(
            'created_by', 'assigned_to', 'client'
        ).prefetch_related(
            'comments', 'status_history', 'attachments'
        )

        # For retrieve, update, destroy, and custom detail actions, return all transactions
        # and let check_object_permissions handle the access control
        if self.action in ['retrieve', 'update', 'partial_update', 'destroy', 'history', 'comments', 'qr_code']:
            return queryset

        # For list action, filter based on user role
        if user.role == 'admin':
            # Admins can see all transactions
            return queryset
        elif user.role == 'editor':
            # Editors can see all transactions they're assigned to or created
            return queryset.filter(
                Q(assigned_to=user) | Q(created_by=user)
            )
        elif user.role == 'client':
            # Clients can only see their own transactions
            return queryset.filter(client=user)

        return Transaction.objects.none()
    
    def check_object_permissions(self, request, obj):
        """
        Check if user has permission to access this specific transaction
        """
        user = request.user

        # Admin has access to everything
        if user.role == 'admin':
            return

        # Editor has access to assigned or created transactions
        elif user.role == 'editor':
            if obj.assigned_to != user and obj.created_by != user:
                self.permission_denied(request, message="You don't have access to this transaction")

        # Client has access only to their own transactions
        elif user.role == 'client':
            if obj.client != user:
                self.permission_denied(request, message="You don't have access to this transaction")

        # Additional check for modification permissions
        if request.method not in permissions.SAFE_METHODS:
            if user.role == 'client' and obj.status not in ['draft']:
                self.permission_denied(request, message="You can only modify draft transactions")
            elif user.role == 'editor':
                # Editors can only update their assigned or created transactions
                if obj.assigned_to != user and obj.created_by != user:
                    self.permission_denied(request, message="You can only modify transactions assigned to you or created by you")
    
    def perform_create(self, serializer):
        """
        Save new transaction with proper audit logging
        """
        # Add request user to context for the serializer
        serializer.context['request'] = self.request
        transaction = serializer.save(created_by=self.request.user)
        
        # Create audit log
        create_audit_log_entry(
            user=self.request.user,
            action='create',
            object_type='Transaction',
            object_id=transaction.id,
            details={
                'transaction_id': transaction.transaction_id,
                'client_name': transaction.client_name,
                'status': transaction.status
            }
        )
        
        # Send notification to assigned user
        if transaction.assigned_to:
            self._send_assignment_notification(transaction)
    
    def perform_update(self, serializer):
        """
        Update transaction with audit logging and optimistic locking
        """
        old_instance = self.get_object()

        # Check for concurrent modification (optimistic locking)
        if 'version' in self.request.data:
            client_version = self.request.data.get('version')
            if client_version and int(client_version) != old_instance.version:
                raise serializers.ValidationError({
                    'error': 'Transaction has been modified by another user. Please refresh and try again.'
                })

        # Store request context on user for audit log
        user = self.request.user
        user._ip_address = get_client_ip(self.request)
        user._user_agent = self.request.META.get('HTTP_USER_AGENT', 'MDC-System/1.0')

        transaction = serializer.save()

        # Create audit log
        create_audit_log_entry(
            user=user,
            action='update',
            object_type='Transaction',
            object_id=transaction.id,
            details={
                'transaction_id': transaction.transaction_id,
                'updated_fields': list(serializer.validated_data.keys()),
                'old_status': old_instance.status,
                'new_status': transaction.status
            }
        )

        # Send notification if assignment changed
        if (old_instance.assigned_to != transaction.assigned_to and
            transaction.assigned_to):
            self._send_assignment_notification(transaction)
    
    def perform_destroy(self, instance):
        """
        Soft delete transaction with audit logging
        """
        instance.soft_delete(self.request.user, "Deleted via API")
        
        create_audit_log_entry(
            user=self.request.user,
            action='delete',
            object_type='Transaction',
            object_id=instance.id,
            details={
                'transaction_id': instance.transaction_id,
                'client_name': instance.client_name
            }
        )
    
    def _send_assignment_notification(self, transaction):
        """
        Send notification email for transaction assignment
        """
        try:
            send_notification_email(
                recipient_email=transaction.assigned_to.email,
                subject=f'Transaction Assignment - {transaction.transaction_id}',
                template_name='emails/transaction_assigned.html',
                context={
                    'transaction': transaction,
                    'assignee': transaction.assigned_to,
                    'frontend_url': settings.MDC_SETTINGS['FRONTEND_URL']
                }
            )
        except Exception as e:
            logger.error(f"Failed to send assignment notification: {str(e)}")
    
    @action(detail=False, methods=['post'], permission_classes=[IsActiveUser])
    def search(self, request):
        """
        Advanced search for transactions with filters
        """
        queryset = self.get_queryset()
        
        # Get search parameters
        search_params = request.data
        
        # Text search
        if search_params.get('query'):
            query = search_params['query']
            queryset = queryset.filter(
                Q(transaction_id__icontains=query) |
                Q(reference_number__icontains=query) |
                Q(client_name__icontains=query) |
                Q(description__icontains=query) |
                Q(tags__icontains=query)
            )
        
        # Status filter
        if search_params.get('status'):
            if isinstance(search_params['status'], list):
                queryset = queryset.filter(status__in=search_params['status'])
            else:
                queryset = queryset.filter(status=search_params['status'])
        
        # Priority filter
        if search_params.get('priority'):
            if isinstance(search_params['priority'], list):
                queryset = queryset.filter(priority__in=search_params['priority'])
            else:
                queryset = queryset.filter(priority=search_params['priority'])
        
        # Category filter
        # Category field removed
        
        # Date range filter
        if search_params.get('date_from'):
            queryset = queryset.filter(created_at__gte=search_params['date_from'])
        
        if search_params.get('date_to'):
            queryset = queryset.filter(created_at__lte=search_params['date_to'])
        
        # Due date filter
        if search_params.get('due_date_from'):
            queryset = queryset.filter(due_date__gte=search_params['due_date_from'])
        
        if search_params.get('due_date_to'):
            queryset = queryset.filter(due_date__lte=search_params['due_date_to'])
        
        # Overdue filter
        if search_params.get('overdue'):
            queryset = queryset.filter(
                due_date__lt=timezone.now().date(),
                status__in=['submitted', 'under_review', 'approved', 'in_progress']
            )
        
        # Assignment filter
        if search_params.get('assigned_to'):
            queryset = queryset.filter(assigned_to__id=search_params['assigned_to'])
        
        if search_params.get('unassigned'):
            queryset = queryset.filter(assigned_to__isnull=True)
        
        # Client filter
        if search_params.get('client'):
            queryset = queryset.filter(client__id=search_params['client'])
        
        # Sorting
        sort_by = search_params.get('sort_by', '-created_at')
        queryset = queryset.order_by(sort_by)
        
        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = TransactionListSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)
        
        serializer = TransactionListSerializer(queryset, many=True, context={'request': request})
        return create_success_response(
            message="Search results retrieved successfully",
            data=serializer.data
        )
    
    def export(self, request):
        """
        Export transactions to Excel/CSV - custom action that returns HttpResponse directly
        """
        # Check permissions manually since we're bypassing DRF
        if not request.user.is_authenticated or not request.user.is_active:
            from django.http import JsonResponse
            return JsonResponse({"success": False, "message": "Authentication required"}, status=401)

        import csv
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        queryset = self.get_queryset()
        
        # Apply filters from query params
        if request.query_params.get('status'):
            queryset = queryset.filter(status=request.query_params.get('status'))
        
        if request.query_params.get('date_from'):
            queryset = queryset.filter(created_at__gte=request.query_params.get('date_from'))
        
        if request.query_params.get('date_to'):
            queryset = queryset.filter(created_at__lte=request.query_params.get('date_to'))
        
        export_format = request.query_params.get('format', 'excel')
        
        if export_format == 'csv':
            # CSV export
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Transaction ID', 'Reference Number', 'Client Name', 
                'Status', 'Priority', 'Category', 'Assigned To',
                'Created At', 'Due Date', 'Description'
            ])
            
            for transaction in queryset:
                writer.writerow([
                    transaction.transaction_id,
                    transaction.reference_number,
                    transaction.client_name,
                    transaction.get_status_display(),
                    transaction.get_priority_display(),
                    '',  # Category field removed
                    transaction.assigned_to.get_full_name() if transaction.assigned_to else '',
                    transaction.created_at.strftime('%Y-%m-%d %H:%M'),
                    transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '',
                    transaction.description
                ])
            
            return response
        
        else:
            # Excel export
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Transactions'
            
            # Headers
            headers = [
                'Transaction ID', 'Reference Number', 'Client Name', 
                'Status', 'Priority', 'Category', 'Assigned To',
                'Created At', 'Due Date', 'Description'
            ]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Data
            for row_num, transaction in enumerate(queryset, 2):
                ws.cell(row=row_num, column=1, value=transaction.transaction_id)
                ws.cell(row=row_num, column=2, value=transaction.reference_number)
                ws.cell(row=row_num, column=3, value=transaction.client_name)
                ws.cell(row=row_num, column=4, value=transaction.get_status_display())
                ws.cell(row=row_num, column=5, value=transaction.get_priority_display())
                ws.cell(row=row_num, column=6, value='')  # Category field removed
                ws.cell(row=row_num, column=7, value=transaction.assigned_to.get_full_name() if transaction.assigned_to else '')
                ws.cell(row=row_num, column=8, value=transaction.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_num, column=9, value=transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '')
                ws.cell(row=row_num, column=10, value=transaction.description)
            
            # Adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
            wb.save(response)
            
            return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsActiveUser])
    def my_transactions(self, request):
        """
        Get current user's transactions
        """
        user = request.user
        
        if user.role == 'client':
            queryset = self.get_queryset().filter(client=user)
        elif user.role == 'editor':
            queryset = self.get_queryset().filter(assigned_to=user)
        else:  # admin
            queryset = self.get_queryset()
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = TransactionListSerializer(page, many=True, context={'request': request})
            return self.get_paginated_response(serializer.data)
        
        serializer = TransactionListSerializer(queryset, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], permission_classes=[IsActiveUser])
    def qr_code(self, request, pk=None):
        """
        Generate QR code for transaction
        """
        transaction = self.get_object()
        
        # Check permissions
        user = request.user
        if user.role == 'client' and transaction.client != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        elif user.role == 'editor' and transaction.assigned_to != user and transaction.created_by != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        try:
            qr_code_base64 = transaction.generate_qr_code()
            return create_success_response(
                message="QR code generated successfully",
                data={
                    'transaction_id': transaction.transaction_id,
                    'qr_code': qr_code_base64
                }
            )
        except Exception as e:
            logger.error(f"Failed to generate QR code: {str(e)}")
            return create_error_response(
                message="Failed to generate QR code",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], permission_classes=[IsActiveUser])
    def history(self, request, pk=None):
        """
        Get transaction history including status changes and activities
        """
        transaction = self.get_object()

        # Check permissions
        user = request.user
        if user.role == 'client' and transaction.client != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        elif user.role == 'editor' and transaction.assigned_to != user and transaction.created_by != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )

        # Get status history
        status_history = transaction.status_history.all().order_by('-created_at')

        # Combine different activity types
        activities = []

        # Add status changes
        for history in status_history:
            activities.append({
                'type': 'status_change',
                'icon': 'bi-arrow-right-circle',
                'color': 'primary',
                'title': f'Status changed from {history.get_previous_status_display() if history.previous_status else "N/A"} to {history.get_new_status_display()}',
                'description': history.change_reason or '',
                'user': history.changed_by.get_full_name() if history.changed_by else 'System',
                'created_at': history.created_at.isoformat(),
            })

        # Add comments as activities
        for comment in transaction.comments.filter(is_deleted=False):
            # Skip internal comments for clients
            if user.role == 'client' and comment.is_internal:
                continue
            activities.append({
                'type': 'comment',
                'icon': 'bi-chat-dots',
                'color': 'info',
                'title': 'Comment added' + (' (Internal)' if comment.is_internal else ''),
                'description': comment.content[:100] + ('...' if len(comment.content) > 100 else ''),
                'user': comment.user.get_full_name() if comment.user else 'Unknown',
                'created_at': comment.created_at.isoformat(),
            })

        # Add attachments as activities
        for attachment in transaction.attachments.filter(is_deleted=False):
            # Skip non-visible attachments for clients
            if user.role == 'client' and not attachment.is_client_visible:
                continue
            activities.append({
                'type': 'attachment',
                'icon': 'bi-paperclip',
                'color': 'success',
                'title': 'File uploaded',
                'description': attachment.original_filename,
                'user': attachment.uploaded_by.get_full_name() if attachment.uploaded_by else 'Unknown',
                'created_at': attachment.created_at.isoformat(),
            })

        # Sort activities by created_at (newest first)
        activities.sort(key=lambda x: x['created_at'], reverse=True)

        return create_success_response(
            message="Transaction history retrieved successfully",
            data={
                'transaction_id': transaction.transaction_id,
                'activities': activities,
                'total_activities': len(activities)
            }
        )

    @action(detail=True, methods=['get', 'post'], permission_classes=[IsActiveUser])
    def attachments(self, request, pk=None):
        """
        Get or upload attachments for a transaction
        """
        transaction = self.get_object()

        if request.method == 'GET':
            # Get all attachments for the transaction
            attachments = Attachment.objects.filter(
                transaction=transaction,
                is_deleted=False
            )

            # Filter based on user role
            if request.user.role == 'client':
                attachments = attachments.filter(is_client_visible=True)

            serializer = AttachmentSerializer(attachments, many=True, context={'request': request})
            return create_success_response(
                message="Attachments retrieved successfully",
                data=serializer.data
            )

        elif request.method == 'POST':
            # Check permission to upload
            if request.user.role == 'client' and transaction.client != request.user:
                return create_error_response(
                    message="You don't have permission to upload attachments to this transaction",
                    status_code=status.HTTP_403_FORBIDDEN
                )

            files = request.FILES.getlist('files')
            if not files:
                return create_error_response(
                    message="No files provided",
                    status_code=status.HTTP_400_BAD_REQUEST
                )

            uploaded_attachments = []
            for file in files:
                attachment = Attachment.objects.create(
                    transaction=transaction,
                    uploaded_by=request.user,
                    file=file,
                    filename=file.name,
                    file_size=file.size,
                    description=request.data.get('description', ''),
                    is_client_visible=request.data.get('is_client_visible', True)
                )
                uploaded_attachments.append(attachment)

            serializer = AttachmentSerializer(
                uploaded_attachments,
                many=True,
                context={'request': request}
            )

            # Log the upload
            logger.info(
                f"User {request.user.username} uploaded {len(files)} files to transaction {transaction.transaction_id}"
            )

            return create_success_response(
                message=f"Successfully uploaded {len(files)} file(s)",
                data=serializer.data,
                status_code=status.HTTP_201_CREATED
            )

    @action(detail=False, methods=['get'], permission_classes=[IsEditorOrAdmin])
    def statistics(self, request):
        """
        Get transaction statistics
        """
        queryset = self.get_queryset()
        
        # Basic counts
        total_transactions = queryset.count()
        by_status = dict(
            queryset.values('status').annotate(count=Count('id')).values_list('status', 'count')
        )
        by_priority = dict(
            queryset.values('priority').annotate(count=Count('id')).values_list('priority', 'count')
        )
        by_category = {}  # Category field removed
        
        # Overdue transactions
        overdue_count = queryset.filter(
            due_date__lt=timezone.now().date(),
            status__in=['submitted', 'under_review', 'approved', 'in_progress']
        ).count()
        
        # Monthly statistics
        current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        completed_this_month = queryset.filter(
            status='completed',
            updated_at__gte=current_month
        ).count()
        
        created_this_month = queryset.filter(
            created_at__gte=current_month
        ).count()
        
        # Average completion time (approximate)
        completed_transactions = queryset.filter(status='completed')
        avg_completion_time = 0
        if completed_transactions.exists():
            avg_delta = completed_transactions.aggregate(
                avg_time=Avg(F('updated_at') - F('created_at'))
            )['avg_time']
            if avg_delta:
                avg_completion_time = avg_delta.days
        
        # Pending assignments
        pending_assignments = queryset.filter(assigned_to__isnull=True).count()
        
        # Top clients
        top_clients = list(
            queryset.values('client__username', 'client__first_name', 'client__last_name')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
        )
        
        # Recent activity (last 7 days)
        recent_date = timezone.now() - timedelta(days=7)
        recent_activity = list(
            queryset.filter(updated_at__gte=recent_date)
            .values('transaction_id', 'client_name', 'status', 'updated_at')
            .order_by('-updated_at')[:10]
        )
        
        stats = {
            'total_transactions': total_transactions,
            'by_status': by_status,
            'by_priority': by_priority,
            'overdue_count': overdue_count,
            'completed_this_month': completed_this_month,
            'created_this_month': created_this_month,
            'avg_completion_time': avg_completion_time,
            'pending_assignments': pending_assignments,
            'top_clients': top_clients,
            'recent_activity': recent_activity
        }
        
        serializer = TransactionStatisticsSerializer(stats)
        return create_success_response(
            message="Transaction statistics retrieved successfully",
            data=serializer.data
        )


class UpdateTransactionStatusView(APIView):
    """
    View for updating transaction status
    """
    permission_classes = [IsActiveUser]
    
    def post(self, request, transaction_id):
        """
        Update transaction status
        """
        try:
            transaction = Transaction.objects.get(
                id=transaction_id,
                is_deleted=False
            )
        except Transaction.DoesNotExist:
            return create_error_response(
                message="Transaction not found",
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        # Check permissions
        user = request.user
        if user.role == 'client' and transaction.client != user:
            return create_error_response(
                message="You don't have permission to modify this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        elif user.role == 'editor' and transaction.assigned_to != user and transaction.created_by != user:
            return create_error_response(
                message="You don't have permission to modify this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        serializer = TransactionStatusUpdateSerializer(
            data=request.data,
            context={'request': request, 'transaction': transaction}
        )
        
        if serializer.is_valid():
            new_status = serializer.validated_data['status']
            reason = serializer.validated_data.get('reason', '')
            
            success = transaction.change_status(new_status, user, reason)
            
            if success:
                # Create audit log
                create_audit_log_entry(
                    user=user,
                    action='status_change',
                    object_type='Transaction',
                    object_id=transaction.id,
                    details={
                        'transaction_id': transaction.transaction_id,
                        'old_status': transaction.status,
                        'new_status': new_status,
                        'reason': reason
                    }
                )
                
                return create_success_response(
                    message="Transaction status updated successfully",
                    data={
                        'transaction_id': transaction.transaction_id,
                        'old_status': transaction.status,
                        'new_status': new_status
                    }
                )
            else:
                return create_error_response(
                    message="Status change not allowed",
                    status_code=status.HTTP_400_BAD_REQUEST
                )
        
        return create_error_response(
            message="Invalid data provided",
            errors=serializer.errors
        )


class AssignTransactionView(APIView):
    """
    View for assigning transactions to users
    """
    permission_classes = [IsEditorOrAdmin]
    
    def post(self, request, transaction_id):
        """
        Assign transaction to user
        """
        try:
            transaction = Transaction.objects.get(
                id=transaction_id,
                is_deleted=False
            )
        except Transaction.DoesNotExist:
            return create_error_response(
                message="Transaction not found",
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        serializer = TransactionAssignmentSerializer(data=request.data)
        
        if serializer.is_valid():
            assigned_to = serializer.validated_data.get('assigned_to')
            reason = serializer.validated_data.get('reason', '')
            
            old_assigned_to = transaction.assigned_to
            transaction.assigned_to = assigned_to
            transaction.save()
            
            # Create audit log
            create_audit_log_entry(
                user=request.user,
                action='assignment_change',
                object_type='Transaction',
                object_id=transaction.id,
                details={
                    'transaction_id': transaction.transaction_id,
                    'old_assigned_to': old_assigned_to.username if old_assigned_to else None,
                    'new_assigned_to': assigned_to.username if assigned_to else None,
                    'reason': reason
                }
            )
            
            # Send notification
            if assigned_to:
                try:
                    send_notification_email(
                        recipient_email=assigned_to.email,
                        subject=f'Transaction Assignment - {transaction.transaction_id}',
                        template_name='emails/transaction_assigned.html',
                        context={
                            'transaction': transaction,
                            'assignee': assigned_to,
                            'frontend_url': settings.MDC_SETTINGS['FRONTEND_URL']
                        }
                    )
                except Exception as e:
                    logger.error(f"Failed to send assignment notification: {str(e)}")
            
            return create_success_response(
                message="Transaction assigned successfully",
                data={
                    'transaction_id': transaction.transaction_id,
                    'assigned_to': assigned_to.username if assigned_to else None
                }
            )
        
        return create_error_response(
            message="Invalid data provided",
            errors=serializer.errors
        )


class TransactionCommentsView(APIView):
    """
    View for managing transaction comments
    """
    permission_classes = [IsActiveUser]
    
    def get(self, request, transaction_id):
        """
        Get transaction comments
        """
        try:
            transaction = get_object_or_404(
                Transaction,
                id=transaction_id,
                is_deleted=False
            )
        except Transaction.DoesNotExist:
            return create_error_response(
                message="Transaction not found",
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        # Check permissions
        user = request.user
        if user.role == 'client' and transaction.client != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        elif user.role == 'editor' and transaction.assigned_to != user and transaction.created_by != user:
            return create_error_response(
                message="You don't have permission to view this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        # Get comments with filtering for clients
        comments = transaction.comments.filter(is_deleted=False).order_by('-created_at')
        if user.role == 'client':
            comments = comments.filter(is_internal=False)
        
        serializer = CommentSerializer(comments, many=True, context={'request': request})
        
        return create_success_response(
            message="Comments retrieved successfully",
            data=serializer.data
        )
    
    def post(self, request, transaction_id):
        """
        Add comment to transaction
        """
        try:
            transaction = get_object_or_404(
                Transaction,
                id=transaction_id,
                is_deleted=False
            )
        except Transaction.DoesNotExist:
            return create_error_response(
                message="Transaction not found",
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        # Check permissions
        user = request.user
        if user.role == 'client' and transaction.client != user:
            return create_error_response(
                message="You don't have permission to comment on this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        elif user.role == 'editor' and transaction.assigned_to != user and transaction.created_by != user:
            return create_error_response(
                message="You don't have permission to comment on this transaction",
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        serializer = CommentSerializer(
            data=request.data,
            context={'request': request, 'transaction': transaction}
        )
        
        if serializer.is_valid():
            comment = serializer.save()
            
            # Create audit log
            create_audit_log_entry(
                user=user,
                action='add_comment',
                object_type='Transaction',
                object_id=transaction.id,
                details={
                    'transaction_id': transaction.transaction_id,
                    'comment_id': comment.id,
                    'is_internal': comment.is_internal
                }
            )
            
            return create_success_response(
                message="Comment added successfully",
                data=serializer.data,
                status_code=status.HTTP_201_CREATED
            )
        
        return create_error_response(
            message="Invalid data provided",
            errors=serializer.errors
        )


class BulkTransactionOperationsView(APIView):
    """
    View for bulk transaction operations
    """
    permission_classes = [IsEditorOrAdmin]
    throttle_classes = [BulkOperationThrottle]
    
    def post(self, request):
        """
        Perform bulk operations on transactions
        """
        serializer = BulkTransactionOperationSerializer(data=request.data)
        
        if serializer.is_valid():
            transaction_ids = serializer.validated_data['transaction_ids']
            action = serializer.validated_data['action']
            
            try:
                with db_transaction.atomic():
                    transactions = Transaction.objects.filter(
                        id__in=transaction_ids,
                        is_deleted=False
                    )
                    updated_count = 0
                    
                    for transaction in transactions:
                        if action == 'assign':
                            assigned_to = serializer.validated_data.get('assigned_to')
                            transaction.assigned_to = assigned_to
                            transaction.save()
                            updated_count += 1
                        
                        elif action == 'change_status':
                            new_status = serializer.validated_data['status']
                            if transaction.can_change_status_to(new_status, request.user):
                                transaction.change_status(
                                    new_status,
                                    request.user,
                                    serializer.validated_data.get('reason', '')
                                )
                                updated_count += 1
                        
                        elif action == 'delete':
                            transaction.soft_delete(
                                request.user,
                                serializer.validated_data.get('reason', '')
                            )
                            updated_count += 1
                        
                        elif action == 'bulk_update':
                            if serializer.validated_data.get('priority'):
                                transaction.priority = serializer.validated_data['priority']
                                transaction.save()
                                updated_count += 1
                    
                    # Create audit log
                    create_audit_log_entry(
                        user=request.user,
                        action=f'bulk_{action}',
                        object_type='Transaction',
                        object_id='bulk',
                        details={
                            'transaction_ids': transaction_ids,
                            'action': action,
                            'updated_count': updated_count,
                            'total_requested': len(transaction_ids)
                        }
                    )
                    
                    return create_success_response(
                        message=f"Bulk {action} completed",
                        data={
                            'total_requested': len(transaction_ids),
                            'updated_count': updated_count
                        }
                    )
                    
            except Exception as e:
                logger.error(f"Error in bulk transaction operation: {str(e)}")
                return create_error_response(
                    message="Error processing bulk operation",
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        return create_error_response(
            message="Invalid data provided",
            errors=serializer.errors
        )


class RecentTransactionsView(APIView):
    """
    Get recent transactions for dashboard
    """
    permission_classes = [IsActiveUser]
    
    def get(self, request):
        """
        Get recent transactions based on user role
        """
        user = request.user
        limit = min(int(request.query_params.get('limit', 10)), 50)
        
        # Get user role with proper error handling
        user_role = getattr(user, 'role', 'client')
        
        # Filter transactions based on role
        if user_role == 'client':
            transactions = Transaction.objects.filter(
                client=user,
                is_deleted=False
            )
        elif user_role == 'editor':
            transactions = Transaction.objects.filter(
                Q(assigned_to=user) | Q(created_by=user),
                is_deleted=False
            )
        else:  # admin
            transactions = Transaction.objects.filter(is_deleted=False)
        
        # Get recent transactions with select_related for better performance
        recent_transactions = transactions.select_related(
            'client', 'assigned_to', 'created_by'
        ).order_by('-created_at')[:limit]
        
        # Serialize the data with context
        serializer = TransactionListSerializer(
            recent_transactions, 
            many=True,
            context={'request': request}
        )
        
        return create_success_response(
            message="Recent transactions retrieved successfully",
            data={
                'results': serializer.data,
                'count': len(serializer.data)
            }
        )
