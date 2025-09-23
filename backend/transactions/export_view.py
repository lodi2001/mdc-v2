"""
Export view for transactions - bypasses DRF JSON rendering
"""
import csv
from django.http import HttpResponse, JsonResponse
from django.views import View
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from .models import Transaction


class ExportTransactionsView(View):
    """Handle transaction exports to Excel/CSV"""

    def get(self, request):
        """Export transactions based on format parameter"""

        # Manual JWT authentication check
        auth = JWTAuthentication()
        try:
            user_auth_tuple = auth.authenticate(request)
            if user_auth_tuple is not None:
                request.user = user_auth_tuple[0]
            else:
                return JsonResponse({"success": False, "message": "Authentication required"}, status=401)
        except Exception as e:
            return JsonResponse({"success": False, "message": "Invalid authentication"}, status=401)

        if not request.user or not request.user.is_authenticated:
            return JsonResponse({"success": False, "message": "Authentication required"}, status=401)

        # Get format from query params
        export_format = request.GET.get('format', 'excel')

        # Get all transactions (you can add filters here later)
        queryset = Transaction.objects.all()

        # Apply user-based filtering
        user = request.user
        if user.role == 'client':
            queryset = queryset.filter(client=user)
        elif user.role == 'editor':
            queryset = queryset.filter(assigned_to=user)

        # Order by created_at
        queryset = queryset.order_by('-created_at')

        if export_format == 'csv':
            # CSV export
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="transactions.csv"'

            writer = csv.writer(response)
            writer.writerow([
                'Transaction ID', 'Reference Number', 'Client Name',
                'Status', 'Priority', 'Category', 'Assigned To',
                'Created At', 'Due Date', 'Description'
            ])

            for transaction in queryset:
                writer.writerow([
                    transaction.transaction_id,
                    transaction.reference_number or '',
                    transaction.client_name,
                    transaction.get_status_display() if hasattr(transaction, 'get_status_display') else transaction.status,
                    transaction.get_priority_display() if hasattr(transaction, 'get_priority_display') else transaction.priority,
                    '',  # Category field removed
                    transaction.assigned_to.get_full_name() if transaction.assigned_to else '',
                    transaction.created_at.strftime('%Y-%m-%d %H:%M'),
                    transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '',
                    transaction.description or ''
                ])

            return response
        else:
            # Excel export
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Transactions'

            # Headers
            headers = [
                'Transaction ID', 'Reference Number', 'Client Name',
                'Status', 'Priority', 'Category', 'Assigned To',
                'Created At', 'Due Date', 'Description'
            ]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)

            # Data
            for row_num, transaction in enumerate(queryset, 2):
                ws.cell(row=row_num, column=1, value=transaction.transaction_id)
                ws.cell(row=row_num, column=2, value=transaction.reference_number or '')
                ws.cell(row=row_num, column=3, value=transaction.client_name)
                ws.cell(row=row_num, column=4, value=transaction.get_status_display() if hasattr(transaction, 'get_status_display') else transaction.status)
                ws.cell(row=row_num, column=5, value=transaction.get_priority_display() if hasattr(transaction, 'get_priority_display') else transaction.priority)
                ws.cell(row=row_num, column=6, value='')  # Category field removed
                ws.cell(row=row_num, column=7, value=transaction.assigned_to.get_full_name() if transaction.assigned_to else '')
                ws.cell(row=row_num, column=8, value=transaction.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_num, column=9, value=transaction.due_date.strftime('%Y-%m-%d') if transaction.due_date else '')
                ws.cell(row=row_num, column=10, value=transaction.description or '')

            # Adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
            wb.save(response)

            return response